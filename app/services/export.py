"""Export scored email data to Excel."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.email import Email
from app.models.score import Score

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial")

SCORE_DIMS = ["personalisation", "clarity", "value_proposition", "cta", "overall"]

EMAIL_SCORES_HEADERS = [
    "Rep", "Subject", "Date", "Personalisation", "Clarity",
    "Value Proposition", "CTA", "Overall", "Notes",
]

REP_AVERAGES_HEADERS = [
    "Rep", "Personalisation", "Clarity", "Value Proposition", "CTA", "Overall",
]


def _score_fill(value: int | None) -> PatternFill | None:
    """Return the fill colour for a score value."""
    if value is None:
        return None
    if value >= 8:
        return GREEN_FILL
    if value >= 6:
        return YELLOW_FILL
    if value >= 4:
        return ORANGE_FILL
    return RED_FILL


async def export_to_excel(session: AsyncSession, output_path: str) -> str:
    """Export scored emails and rep averages to an xlsx workbook.

    Sheet 1 "Email Scores": one row per scored email with colour-coded score cells.
    Sheet 2 "Rep Averages": one row per rep, sorted by overall average descending.

    Returns the output_path.
    """
    wb = Workbook()

    # --- Sheet 1: Email Scores ---
    ws = wb.active
    ws.title = "Email Scores"
    ws.append(EMAIL_SCORES_HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"

    stmt = (
        select(Email, Score)
        .join(Score, Email.id == Score.email_id)
        .where(Score.score_error.is_(False))
    )
    result = await session.execute(stmt)
    rows = result.all()

    for email, score in rows:
        ws.append([
            email.from_name or email.from_email,
            email.subject,
            email.timestamp,
            score.personalisation,
            score.clarity,
            score.value_proposition,
            score.cta,
            score.overall,
            score.notes,
        ])
        row_num = ws.max_row
        for col in range(1, len(EMAIL_SCORES_HEADERS) + 1):
            ws.cell(row=row_num, column=col).font = BODY_FONT
        # Colour-code score columns (4 through 8)
        for i, dim in enumerate(SCORE_DIMS):
            cell = ws.cell(row=row_num, column=4 + i)
            fill = _score_fill(getattr(score, dim))
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions

    # --- Sheet 2: Rep Averages ---
    ws2 = wb.create_sheet("Rep Averages")
    ws2.append(REP_AVERAGES_HEADERS)
    for cell in ws2[1]:
        cell.font = HEADER_FONT
    ws2.freeze_panes = "A2"

    avg_stmt = (
        select(
            Email.from_email,
            func.avg(Score.personalisation).label("avg_personalisation"),
            func.avg(Score.clarity).label("avg_clarity"),
            func.avg(Score.value_proposition).label("avg_value_proposition"),
            func.avg(Score.cta).label("avg_cta"),
            func.avg(Score.overall).label("avg_overall"),
        )
        .join(Score, Email.id == Score.email_id)
        .where(Score.score_error.is_(False))
        .group_by(Email.from_email)
        .order_by(func.avg(Score.overall).desc())
    )
    result = await session.execute(avg_stmt)

    for row in result.all():
        ws2.append([
            row.from_email,
            round(float(row.avg_personalisation), 1),
            round(float(row.avg_clarity), 1),
            round(float(row.avg_value_proposition), 1),
            round(float(row.avg_cta), 1),
            round(float(row.avg_overall), 1),
        ])
        row_num = ws2.max_row
        for col in range(1, len(REP_AVERAGES_HEADERS) + 1):
            ws2.cell(row=row_num, column=col).font = BODY_FONT

    ws2.auto_filter.ref = ws2.dimensions

    wb.save(output_path)
    return output_path


async def export_rep_emails(
    session: AsyncSession,
    rep_email: str,
    *,
    search: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    score_min: int | None = None,
    score_max: int | None = None,
    export_all: bool = False,
    email_type: str | None = None,
) -> BytesIO:
    """Export a single rep's scored emails to an Excel workbook in memory.

    When export_all is True, all filter params are ignored and every scored
    email for the rep is included.

    Returns a BytesIO buffer containing the .xlsx file.
    """
    filters = [
        Email.from_email == rep_email,
        Score.score_error.is_(False),
    ]

    if email_type in ("outreach", "follow_up"):
        from app.services.rep import _follow_up_ids
        fu_ids = await _follow_up_ids(session, rep_email)
        filters.append(Email.chain_id.is_(None))
        if email_type == "outreach":
            if fu_ids:
                filters.append(Email.id.notin_(fu_ids))
        elif email_type == "follow_up":
            if fu_ids:
                filters.append(Email.id.in_(fu_ids))
            else:
                filters.append(Email.id == -1)

    if not export_all:
        if search:
            pattern = f"%{search}%"
            filters.append(
                or_(
                    Email.subject.ilike(pattern),
                    Email.body_text.ilike(pattern),
                )
            )
        if date_from:
            filters.append(Email.timestamp >= date_from)
        if date_to:
            filters.append(Email.timestamp <= date_to)
        if score_min is not None:
            filters.append(Score.overall >= score_min)
        if score_max is not None:
            filters.append(Score.overall <= score_max)

    stmt = (
        select(Email, Score)
        .join(Score, Email.id == Score.email_id)
        .where(*filters)
        .order_by(Email.timestamp.desc())
    )
    result = await session.execute(stmt)
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Email Scores"
    ws.append(EMAIL_SCORES_HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"

    for email, score in rows:
        ws.append([
            email.from_name or email.from_email,
            email.subject,
            email.timestamp,
            score.personalisation,
            score.clarity,
            score.value_proposition,
            score.cta,
            score.overall,
            score.notes,
        ])
        row_num = ws.max_row
        for col in range(1, len(EMAIL_SCORES_HEADERS) + 1):
            ws.cell(row=row_num, column=col).font = BODY_FONT
        for i, dim in enumerate(SCORE_DIMS):
            cell = ws.cell(row=row_num, column=4 + i)
            fill = _score_fill(getattr(score, dim))
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


CHAIN_HEADERS = [
    "Subject", "Emails", "Last Activity",
    "Progression", "Responsiveness", "Persistence",
    "Quality", "Notes",
]

CHAIN_SCORE_DIMS = [
    "progression", "responsiveness", "persistence", "conversation_quality",
]


async def export_rep_chains(
    session: AsyncSession,
    rep_email: str,
    *,
    search: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    score_min: int | None = None,
    score_max: int | None = None,
    export_all: bool = False,
    status: str | None = None,
) -> BytesIO:
    """Export a rep's conversation chains to an Excel workbook in memory."""
    from app.services.chain import get_rep_chains

    if export_all:
        result = await get_rep_chains(session, rep_email, page=1, per_page=10000)
    else:
        result = await get_rep_chains(
            session, rep_email, page=1, per_page=10000,
            search=search, date_from=date_from, date_to=date_to,
            score_min=score_min, score_max=score_max, status=status,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Conversations"
    ws.append(CHAIN_HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"

    for chain in result["items"]:
        ws.append([
            chain.get("normalized_subject", ""),
            chain.get("email_count", 0),
            chain.get("last_activity_at"),
            chain.get("progression"),
            chain.get("responsiveness"),
            chain.get("persistence"),
            chain.get("conversation_quality"),
            "",
        ])
        row_num = ws.max_row
        for col in range(1, len(CHAIN_HEADERS) + 1):
            ws.cell(row=row_num, column=col).font = BODY_FONT
        for i, dim in enumerate(CHAIN_SCORE_DIMS):
            cell = ws.cell(row=row_num, column=4 + i)
            fill = _score_fill(chain.get(dim))
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
